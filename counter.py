import requests
import tkinter as tk
from tkinter import *
from time import sleep
from threading import Thread
from openpyxl import load_workbook

from tkinter import ttk


import pandas as pd
import xlsxwriter
miain_count =0
flag = True
flag2 = True


dic ={}

count=0
Calbri_font1 = ('Calibri', 13, 'bold')
Calbri_font2 = ('Calibri', 11, 'normal')

requests.packages.urllib3.disable_warnings()

event_nameset =["private1"]
event_nameset1 =["gold","mpls","private1","default","biz-internet","public-internet"]
site_ip =[]
pvt_ip =[]
mpl_ip =[]
mpls_bdf_up_clear=[]
pvt_bdf_up_clear=[]
SdwanName=[]
MplsID=[]
BizDslID=[]
PvtID=[]
global q
previous_mpls=""
previous_pvt=""
m=0
wheight=0

     



def dataID_find(search_ip):

    try:
        wb = load_workbook('data.xlsx')
        ws = wb.active
        col_A = ws['A']
        count = 0
        
        find_flag = False
        for cell in col_A:
            #print(cell.value)
            count+=1
            if cell.value == search_ip:
                find_flag = True
                break       
                
        if find_flag:
            
            dialog_id =(ws.cell(row=(count),column=4).value)
            mpls_id = (ws.cell(row=(count),column=2).value)
            bizdsl_id = (ws.cell(row=(count),column=3).value)
            sdwn_mname = (ws.cell(row=(count),column=5).value)
            find_flag = False
            return(sdwn_mname,mpls_id,bizdsl_id,dialog_id)
        else:
            #return('no matching','no matching')
            return("None","None","None","None")
    except Exception as e:
        print(e)




    #command=config()
def config(name,mp,pv,mpls_color,pvt_color,pos_x,pos_y):
        global count
       
            
        fram1= tk.Frame(second_frame,background='#03506F')
        fram1.grid(row =pos_x,column=pos_y)


        fram= tk.Frame(fram1,padx=1,pady=1,background='#0A043C')
        fram.grid(padx=4,pady=5)
    
    #fram.grid_propagate(0)
    

        aa = tk.Label(fram,text=name,foreground="black",background="gold",padx=12,width=20,font=Calbri_font1,relief="raised")
        aa.pack(pady=(0,7))
    #num.pack_propagate(False)


        bb= tk.Label(fram,text=f'SLT:  {mp}',foreground="black",background=pvt_color,width=19,font=Calbri_font2,relief="groove")
        bb.pack(pady=(0,5))

        cc = tk.Label(fram,text=f'Dialog:  {pv}',foreground="black",background=mpls_color,width=19,font=Calbri_font2,relief="groove")
        cc.pack(pady=(0,8))




def api_call():
    
    print("api")
    USERNAME="niranga_slt"
    PASSWORD="N1r@ng@#"

    site_ip.clear()
    pvt_ip.clear()
    mpl_ip.clear()
    mpls_bdf_up_clear.clear()
    pvt_bdf_up_clear.clear()
    #SdwanName.clear()
    #MplsID.clear()
    #BizDslID.clear()
    #PvtID.clear()

    try:
        login_url = 'https://vmanage-2203971.sdwan.cisco.com/j_security_check'
        login_data = {'j_username' : USERNAME, 'j_password' : PASSWORD}
        response = requests.post(url=login_url, data=login_data, verify=False,timeout=20)
        if response.content != b'':
            print ("Authentication fail!")
            exit()
        token1='JSESSIONID='+response.cookies.get_dict()['JSESSIONID']
        print ('token1=',token1)
        token_url='https://vmanage-2203971.sdwan.cisco.com/dataservice/client/token'
        headers = {'Cookie': token1,'content-type':'application/json'}
        response = requests.get(url=token_url, headers=headers, verify=False,timeout=20)
        token2=response.text
        print ('token2=',token2)
        ip = '1.6.1.47'
        headers = {'Content-Type': "application/json",'Cookie': token1, 'X-XSRF-TOKEN': token2}
        request_url='https://vmanage-2203971.sdwan.cisco.com/dataservice/device/tloc'
        response = requests.get(url=request_url,headers=headers,verify=False,timeout=20)
        #print ("status:",response.text)
        if response.status_code ==200:
            jstring = response.json()
            count =0
            for mpl in jstring["data"] :
                if mpl["color"]== "mpls":
                    
                    #print(mpl["system-ip"])
                    mlps_id = mpl["system-ip"]
                    mpl_ip.append(mpl["system-ip"])
                    mpls_bdf_up_clear.append(mpl["bfdSessionsUp"])
                    #mlps_id = mpl["system-ip"]
                
                    for mpl in jstring["data"] :    
                        if mpl["system-ip"]==mlps_id:
                            if mpl["color"] in event_nameset:
                                pvt_bdf_up_clear.append(mpl["bfdSessionsUp"]) 
                                pvt_ip.append(mlps_id)

              

                                        
                    
            len1 = len(mpl_ip)                 
            print(len1)
            f = lambda list1,list2:list(filter(lambda element: element not in list2,list1 ))
            union=(f(mpl_ip,pvt_ip))
            print(union)


            #print(len(mpl_ip))
            #print(len(mpls_bdf_up_clear))

            #print(len(pvt_bdf_up_clear))
            #print(len(pvt_ip))
            
            for ips in union:
                pvt_bdf_up_clear.insert(mpl_ip.index(ips),'bumy')
            
            #print(len(mpl_ip))
            #print(len(mpls_bdf_up_clear))

            #print(len(pvt_bdf_up_clear))
            #print(len(pvt_ip))

            #print((mpl_ip))
            #print((pvt_ip))
            global miain_count
            if miain_count == 0:
                for ids in mpl_ip:
                    x,y,z,a=dataID_find(ids)
                    SdwanName.append(x)
                    MplsID.append(y)
                    BizDslID.append(z)
                    PvtID.append(a)

            print((mpls_bdf_up_clear))

            print((pvt_bdf_up_clear))

            global previous_mpls
            global previous_pvt


            print(len(previous_mpls))
            print(len(previous_pvt))
            

            miain_count+=1
            print(f"counter{miain_count}") 
            

            
                        

            if (previous_mpls == mpls_bdf_up_clear) and (previous_pvt == pvt_bdf_up_clear):
                print("no chanage in MPLS")
                global w
                print("noch:", w)
                #config("abc","green2","red2",0,0)
            else:
                pass
                print("chanage in MPLS or PVT")   
                w=1
                print("ch:", w)
                for index_m,(old_m,new_m) in enumerate(zip(previous_mpls,mpls_bdf_up_clear)):
                    if(old_m!=new_m):
                        print(mpl_ip[index_m],new_m)

                        if (new_m ==0)and (pvt_bdf_up_clear[mpl_ip.index(mpl_ip[index_m])])=='bumy':
                            #print(pvt_bdf_up_clear[mpl_ip.index(mpl_ip[index_m])])
                            x,y,_,a=dataID_find(mpl_ip[index_m])
                            config(x,y,a,"cyan","red2",dic.get(mpl_ip[index_m])[0],dic.get(mpl_ip[index_m])[1])

                        elif (new_m ==0) and (pvt_bdf_up_clear[mpl_ip.index(mpl_ip[index_m])]==0):
                                x,y,_,a=dataID_find(mpl_ip[index_m])
                                config(x,y,a,"red2","red2",dic.get(mpl_ip[index_m])[0],dic.get(mpl_ip[index_m])[1])

                        elif (new_m ==0) and (pvt_bdf_up_clear[mpl_ip.index(mpl_ip[index_m])]>0):
                                x,y,_,a=dataID_find(mpl_ip[index_m])

                                config(x,y,a,"green2","red2",dic.get(mpl_ip[index_m])[0],dic.get(mpl_ip[index_m])[1])

                        elif (new_m >0)and (pvt_bdf_up_clear[mpl_ip.index(mpl_ip[index_m])])=='bumy':
                            #print(pvt_bdf_up_clear[mpl_ip.index(mpl_ip[index_m])])
                            x,y,_,a=dataID_find(mpl_ip[index_m])
                            config(x,y,a,"cyan","green2",dic.get(mpl_ip[index_m])[0],dic.get(mpl_ip[index_m])[1])

                        elif (new_m >0) and (pvt_bdf_up_clear[mpl_ip.index(mpl_ip[index_m])]==0):
                                x,y,_,a=dataID_find(mpl_ip[index_m])
                                config(x,y,a,"red2","green2",dic.get(mpl_ip[index_m])[0],dic.get(mpl_ip[index_m])[1])

                        elif (new_m >0) and (pvt_bdf_up_clear[mpl_ip.index(mpl_ip[index_m])]>0):
                                x,y,_,a=dataID_find(mpl_ip[index_m])
                                config(x,y,a,"green2","green2",dic.get(mpl_ip[index_m])[0],dic.get(mpl_ip[index_m])[1])


                for index_p,(old_p,new_p) in enumerate(zip(previous_pvt,pvt_bdf_up_clear)):
                        if(old_p!=new_p):
                            print(mpl_ip[index_p],new_p)

                            
                            if (new_p ==0) and (mpls_bdf_up_clear[mpl_ip.index(mpl_ip[index_p])]==0):
                                    x,y,_,a=dataID_find(mpl_ip[index_p])
                                    config(x,y,a,"red2","red2",dic.get(mpl_ip[index_p])[0],dic.get(mpl_ip[index_p])[1])

                            elif (new_p ==0) and (mpls_bdf_up_clear[mpl_ip.index(mpl_ip[index_p])]>0):
                                    x,y,_,a=dataID_find(mpl_ip[index_p])

                                    config(x,y,a,"red2","green2",dic.get(mpl_ip[index_p])[0],dic.get(mpl_ip[index_p])[1])
                                    
                            
                            elif (new_p >0) and (mpls_bdf_up_clear[mpl_ip.index(mpl_ip[index_p])]==0):
                                    x,y,_,a=dataID_find(mpl_ip[index_p])
                                    config(x,y,a,"green2","red2",dic.get(mpl_ip[index_p])[0],dic.get(mpl_ip[index_p])[1])

                            elif (new_p >0) and (mpls_bdf_up_clear[mpl_ip.index(mpl_ip[index_p])]>0):
                                    x,y,_,a=dataID_find(mpl_ip[index_p])
                                    config(x,y,a,"green2","green2",dic.get(mpl_ip[index_p])[0],dic.get(mpl_ip[index_p])[1]) 

                        #print(dic.get(mpl_ip[index])[0])
                        #print(dic.get(mpl_ip[index])[1])
                        

            previous_mpls=list(mpls_bdf_up_clear)
            previous_pvt=list(pvt_bdf_up_clear)
            
            
         
                    
                
      #config(name,mpls_color,pvt_color,pos_x,pos_y):
                #print(dic.get(ids))

                #dic.get(ids)[0]


            
             

            if miain_count==1:
                r = 0
                c = 0

                window = tk.Tk()
                window.geometry("400x400")
                window['background'] = 'green'
                window.configure(bg='red')
                window.columnconfigure(0,weight=1)
                window.rowconfigure(0,weight=1)

                my_canvas = Canvas(window)
                
                
                my_canvas.pack(side=LEFT,fill=BOTH,expand=1)

                my_scrollbar_x =ttk.Scrollbar(window,orient=HORIZONTAL,command=my_canvas.xview)
                my_scrollbar_x.place(relx=0,rely=1,relwidth=1,anchor='sw')
                my_canvas.configure(xscrollcommand=my_scrollbar_x.set)

                my_scrollbar =ttk.Scrollbar(window,orient=VERTICAL,command=my_canvas.yview)
                my_scrollbar.place(relx=1,rely=0,relheight=1,anchor='ne')
                my_canvas.configure(yscrollcommand=my_scrollbar.set)


                my_canvas.bind('<Configure>',lambda e: my_canvas.configure(scrollregion=(my_canvas.bbox("all"))))#my_canvas.bbox("all")
                global second_frame
                second_frame = Frame(my_canvas)
                my_canvas.create_window((0,0),window=second_frame,anchor="nw")#
                second_frame.configure(bg='#03506F')

                

                for num in range(len(mpl_ip)):

                    fram1= tk.Frame(second_frame,background='#03506F')
                    fram1.grid(row =r,column=c)


                    fram= tk.Frame(fram1,padx=1,pady=1,background='#0A043C')
                    fram.grid(padx=4,pady=5)
                    
                    aa = tk.Label(fram,text=SdwanName[num],foreground="black",background="gold",padx=12,width=20,font=Calbri_font1,relief="raised")
                    aa.pack(pady=(0,7))

                    if (mpls_bdf_up_clear[num]!=0):
                        
                        bb= tk.Label(fram,text=f'SLT:  {MplsID[num]}',foreground="black",background="green2",width=19,font=Calbri_font2,relief="groove")
                        bb.pack(pady=(0,5))
                    
                    elif (mpls_bdf_up_clear[num]==0):
                         
                        bb= tk.Label(fram,text=f'SLT:  {MplsID[num]}',foreground="black",background="red2",width=19,font=Calbri_font2,relief="groove")
                        bb.pack(pady=(0,5))

                        
                    if(pvt_bdf_up_clear[num]=='bumy'):
                        

                        cc = tk.Label(fram,text=f'Dialog:  {PvtID[num]}',foreground="black",background="cyan",width=19,font=Calbri_font2,relief="groove")
                        cc.pack(pady=(0,8))

                    elif(pvt_bdf_up_clear[num]==0):
                        

                        cc = tk.Label(fram,text=f'Dialog:  {PvtID[num]}',foreground="black",background="red2",width=19,font=Calbri_font2,relief="groove")
                        cc.pack(pady=(0,8))

                    elif(pvt_bdf_up_clear[num]!=0) and (pvt_bdf_up_clear[num]!='bumy') :
                        

                        cc = tk.Label(fram,text=f'Dailog:  {PvtID[num]}',foreground="black",background="green2",width=19,font=Calbri_font2,relief="groove")
                        cc.pack(pady=(0,8))

                    dic[mpl_ip[num]] =r,c

                    

                    r +=1

                    if r==7:
                        r=0
                        c+=1
                #print(dic)
                th = Thread(target= abc)
                th.daemon = True
                th.start()
                window.mainloop()
            

    except Exception as e:
        print(e)

def abc():
    while True:
        api_call()
        sleep(60) 


if flag2:
    abc()
    flag2=False


 


            












    

    



    



   


          